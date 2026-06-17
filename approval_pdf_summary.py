from __future__ import annotations

import argparse
import json
import os
import re
import shutil
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


PROJECT_ROOT = Path(__file__).resolve().parent
DEFAULT_ENV_PATH = PROJECT_ROOT / ".approval_pdf_summary.env"

DEFAULT_PDF_PREFIX = "审批单_"
DEFAULT_STATUS_TEXT = "有审批单"
DEFAULT_ID_HEADER = "设计变更编号"
DEFAULT_STATUS_HEADER = "审批单情况"

INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
TRAILING_NUMBER_RE = re.compile(r"^(.+)-(\d+)$")


@dataclass(frozen=True)
class PdfRecord:
    change_id: str
    group_name: str
    file_name: str
    file_path: Path
    file_size: int
    modified_at: datetime


def parse_env_file(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}

    values: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[len("export ") :].lstrip()
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]
        if key:
            values[key] = value
    return values


def env_bool(values: dict[str, str], key: str, default: bool) -> bool:
    raw = os.environ.get(key, values.get(key))
    if raw is None or raw == "":
        return default
    return raw.strip().lower() in {"1", "true", "yes", "y", "on"}


def env_int(values: dict[str, str], key: str, default: int) -> int:
    raw = os.environ.get(key, values.get(key))
    if raw is None or raw == "":
        return default
    return int(raw)


def env_path(values: dict[str, str], key: str) -> Path:
    raw = os.environ.get(key, values.get(key))
    if not raw:
        raise ValueError(f"Missing required config: {key}")
    return Path(raw).expanduser()


def natural_key(value: str) -> list[object]:
    parts: list[object] = []
    for part in re.split(r"(\d+)", value):
        if part.isdigit():
            parts.append(int(part))
        else:
            parts.append(part.casefold())
    return parts


def get_group_name(change_id: str) -> str:
    match = TRAILING_NUMBER_RE.match(change_id)
    if match:
        return match.group(1)
    return change_id


def scan_pdfs(pdf_dir: Path, pdf_prefix: str, recursive: bool) -> list[PdfRecord]:
    pattern = "**/*.pdf" if recursive else "*.pdf"
    records: list[PdfRecord] = []
    for file_path in pdf_dir.glob(pattern):
        if not file_path.is_file():
            continue
        stem = file_path.stem.strip()
        if pdf_prefix and stem.startswith(pdf_prefix):
            stem = stem[len(pdf_prefix) :].strip()
        if not stem:
            continue

        stat = file_path.stat()
        records.append(
            PdfRecord(
                change_id=stem,
                group_name=get_group_name(stem),
                file_name=file_path.name,
                file_path=file_path,
                file_size=stat.st_size,
                modified_at=datetime.fromtimestamp(stat.st_mtime),
            )
        )

    return sorted(records, key=lambda item: (natural_key(item.group_name), natural_key(item.change_id), item.file_name))


def safe_sheet_name(name: str, existing_names: Iterable[str]) -> str:
    cleaned = INVALID_SHEET_CHARS.sub("-", name).strip("'").strip() or "审批单"
    cleaned = cleaned[:31]
    if cleaned not in existing_names:
        return cleaned

    base = cleaned[:28]
    index = 2
    used = set(existing_names)
    while True:
        candidate = f"{base}_{index}"[:31]
        if candidate not in used:
            return candidate
        index += 1


def create_workbook(path: Path, records: list[PdfRecord], status_text: str) -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    groups = sorted({record.group_name for record in records}, key=natural_key)
    if not groups:
        groups = ["审批单"]

    for group in groups:
        ws = wb.create_sheet(safe_sheet_name(group, wb.sheetnames))
        ws.cell(1, 1).value = DEFAULT_ID_HEADER
        ws.cell(1, 2).value = DEFAULT_STATUS_HEADER
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 16
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    return wb


def sheet_for_group(wb: Workbook, group_name: str) -> Worksheet:
    if group_name in wb.sheetnames:
        return wb[group_name]

    for ws in wb.worksheets:
        if ws.title.casefold() == group_name.casefold():
            return ws

    return wb.create_sheet(safe_sheet_name(group_name, wb.sheetnames))


def existing_row_by_id(ws: Worksheet, id_col: int, header_row: int) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        raw_value = ws.cell(row, id_col).value
        if raw_value is None:
            continue
        change_id = str(raw_value).strip()
        if change_id and change_id not in rows:
            rows[change_id] = row
    return rows


def ensure_headers(ws: Worksheet, header_row: int, id_col: int, status_col: int) -> None:
    if ws.cell(header_row, id_col).value in (None, ""):
        ws.cell(header_row, id_col).value = DEFAULT_ID_HEADER
    if ws.cell(header_row, status_col).value in (None, ""):
        ws.cell(header_row, status_col).value = DEFAULT_STATUS_HEADER


def copy_row_format(ws: Worksheet, source_row: int, target_row: int, max_col: int) -> None:
    if source_row < 1:
        return

    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
        ws.row_dimensions[target_row].hidden = ws.row_dimensions[source_row].hidden

    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def write_records(
    wb: Workbook,
    records: list[PdfRecord],
    *,
    header_row: int,
    id_col: int,
    status_col: int,
    status_text: str,
    append_missing: bool,
    clear_missing: bool,
    dry_run: bool,
) -> dict[str, object]:
    records_by_group: dict[str, list[PdfRecord]] = {}
    for record in records:
        records_by_group.setdefault(record.group_name, []).append(record)

    updated = 0
    appended = 0
    already_marked = 0
    cleared = 0
    unmatched: list[str] = []
    sheet_summaries: dict[str, dict[str, int]] = {}

    for group_name, group_records in records_by_group.items():
        ws = sheet_for_group(wb, group_name)
        ensure_headers(ws, header_row, id_col, status_col)
        row_lookup = existing_row_by_id(ws, id_col, header_row)
        seen_ids = {record.change_id for record in group_records}
        sheet_updates = 0
        sheet_appends = 0

        for record in group_records:
            row = row_lookup.get(record.change_id)
            if row is None:
                if not append_missing:
                    unmatched.append(record.change_id)
                    continue
                row = ws.max_row + 1
                if not dry_run:
                    copy_row_format(ws, max(header_row, row - 1), row, max(ws.max_column, status_col))
                    ws.cell(row, id_col).value = record.change_id
                    ws.cell(row, status_col).value = status_text
                row_lookup[record.change_id] = row
                appended += 1
                sheet_appends += 1
                continue

            current = ws.cell(row, status_col).value
            if current == status_text:
                already_marked += 1
                continue
            if not dry_run:
                ws.cell(row, status_col).value = status_text
            updated += 1
            sheet_updates += 1

        if clear_missing:
            for change_id, row in row_lookup.items():
                if change_id in seen_ids:
                    continue
                current = ws.cell(row, status_col).value
                if current in (None, ""):
                    continue
                if not dry_run:
                    ws.cell(row, status_col).value = None
                cleared += 1

        sheet_summaries[ws.title] = {
            "pdf_count": len(group_records),
            "updated": sheet_updates,
            "appended": sheet_appends,
        }

    return {
        "pdf_count": len(records),
        "sheet_count": len(wb.sheetnames),
        "updated": updated,
        "appended": appended,
        "already_marked": already_marked,
        "cleared": cleared,
        "unmatched_count": len(unmatched),
        "unmatched": unmatched[:30],
        "sheets": sheet_summaries,
    }


def backup_workbook(path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = path.with_name(f"{path.stem}.backup_{timestamp}{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def run(args: argparse.Namespace) -> dict[str, object]:
    env_values = parse_env_file(Path(args.env))

    pdf_dir = Path(args.pdf_dir) if args.pdf_dir else env_path(env_values, "APPROVAL_PDF_DIR")
    excel_path = Path(args.excel) if args.excel else env_path(env_values, "APPROVAL_EXCEL_PATH")
    pdf_prefix = args.pdf_prefix if args.pdf_prefix is not None else os.environ.get(
        "APPROVAL_PDF_PREFIX", env_values.get("APPROVAL_PDF_PREFIX", DEFAULT_PDF_PREFIX)
    )
    status_text = args.status_text if args.status_text is not None else os.environ.get(
        "APPROVAL_STATUS_TEXT", env_values.get("APPROVAL_STATUS_TEXT", DEFAULT_STATUS_TEXT)
    )

    recursive = args.recursive or env_bool(env_values, "APPROVAL_SCAN_RECURSIVE", False)
    append_missing = (not args.no_append_missing) and env_bool(env_values, "APPROVAL_APPEND_MISSING", True)
    clear_missing = args.clear_missing or env_bool(env_values, "APPROVAL_CLEAR_MISSING", False)
    header_row = env_int(env_values, "APPROVAL_HEADER_ROW", 1)
    id_col = env_int(env_values, "APPROVAL_ID_COLUMN", 1)
    status_col = env_int(env_values, "APPROVAL_STATUS_COLUMN", 2)

    if not pdf_dir.exists() or not pdf_dir.is_dir():
        raise FileNotFoundError(f"PDF directory does not exist: {pdf_dir}")

    records = scan_pdfs(pdf_dir, pdf_prefix, recursive)
    if excel_path.exists():
        wb = load_workbook(excel_path)
        created_workbook = False
    else:
        wb = create_workbook(excel_path, records, status_text)
        created_workbook = True

    before_sheets = wb.sheetnames[:]
    result = write_records(
        wb,
        records,
        header_row=header_row,
        id_col=id_col,
        status_col=status_col,
        status_text=status_text,
        append_missing=append_missing,
        clear_missing=clear_missing,
        dry_run=args.dry_run,
    )

    changed = bool(created_workbook or result["updated"] or result["appended"] or result["cleared"])
    backup_path = None
    saved = False
    if not args.dry_run and changed:
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        if excel_path.exists() and not args.no_backup:
            backup_path = backup_workbook(excel_path)
        wb.save(excel_path)
        saved = True

    result.update(
        {
            "dry_run": args.dry_run,
            "changed": changed,
            "saved": saved,
            "created_workbook": created_workbook,
            "pdf_dir": str(pdf_dir),
            "excel_path": str(excel_path),
            "backup_path": str(backup_path) if backup_path else None,
            "sheet_names_before": before_sheets,
            "sheet_names_after": wb.sheetnames,
        }
    )
    return result


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Scan approval PDF files and summarize their status into an Excel workbook."
    )
    parser.add_argument("--env", default=str(DEFAULT_ENV_PATH), help="Path to the local .env file.")
    parser.add_argument("--pdf-dir", help="Override APPROVAL_PDF_DIR from the .env file.")
    parser.add_argument("--excel", help="Override APPROVAL_EXCEL_PATH from the .env file.")
    parser.add_argument("--pdf-prefix", help="Filename prefix to strip before extracting the change id.")
    parser.add_argument("--status-text", help="Status text written to the approval-status column.")
    parser.add_argument("--recursive", action="store_true", help="Scan PDFs recursively.")
    parser.add_argument("--dry-run", action="store_true", help="Print the planned changes without saving.")
    parser.add_argument("--no-backup", action="store_true", help="Do not create a timestamped backup before saving.")
    parser.add_argument("--no-append-missing", action="store_true", help="Do not append PDF ids missing from Excel.")
    parser.add_argument("--clear-missing", action="store_true", help="Clear existing status cells when the matching PDF is missing.")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    result = run(args)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
