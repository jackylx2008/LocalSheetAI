from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
import warnings
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from localai.modules.config_loader import as_bool, as_int
from localai.modules.excel_reader import resolve_sheet_name


@dataclass(frozen=True)
class ExcelIssueExtractConfig:
    header_row: int = 3
    data_start_row: int = 4
    problem_column: str = "E"
    data_only: bool = True
    max_cell_chars: int = 500

    @classmethod
    def from_config(cls, raw: dict[str, Any]) -> "ExcelIssueExtractConfig":
        return cls(
            header_row=as_int(raw.get("header_row"), 3),
            data_start_row=as_int(raw.get("data_start_row"), 4),
            problem_column=str(raw.get("problem_column", "E")).strip() or "E",
            data_only=as_bool(raw.get("data_only", True)),
            max_cell_chars=as_int(raw.get("max_cell_chars"), 500),
        )


@dataclass(frozen=True)
class ExcelIssueRow:
    workbook_path: Path
    sheet_name: str
    row_number: int
    problem_text: str
    row_context: dict[str, str]
    missing_problem_description: bool


def extract_issue_rows(
    workbook_path: Path,
    sheet_name: str,
    config: ExcelIssueExtractConfig,
) -> list[ExcelIssueRow]:
    sheet_names = _list_sheet_names(workbook_path)
    actual_sheet_name = resolve_sheet_name(sheet_name, sheet_names)
    problem_col_index = column_index_from_string(config.problem_column.upper())

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(
            workbook_path,
            read_only=True,
            data_only=config.data_only,
            keep_links=False,
        )
    try:
        worksheet = workbook[actual_sheet_name]
        max_column = max(worksheet.max_column or 0, problem_col_index)
        headers = _read_headers(worksheet, config.header_row, max_column, config.max_cell_chars)
        rows: list[ExcelIssueRow] = []

        first_data_row = max(config.data_start_row, config.header_row + 1)
        for row_number, values in enumerate(worksheet.iter_rows(
            min_row=first_data_row,
            max_col=max_column,
            values_only=True,
        ), start=first_data_row):
            normalized = [_format_cell(value, config.max_cell_chars) for value in values]
            if not any(normalized):
                continue

            problem_text = normalized[problem_col_index - 1] if problem_col_index <= len(normalized) else ""
            row_context = {
                headers[index]: value
                for index, value in enumerate(normalized)
                if index < len(headers) and value
            }
            rows.append(
                ExcelIssueRow(
                    workbook_path=workbook_path,
                    sheet_name=actual_sheet_name,
                    row_number=row_number,
                    problem_text=problem_text,
                    row_context=row_context,
                    missing_problem_description=not bool(problem_text),
                )
            )

        return rows
    finally:
        workbook.close()


def _read_headers(worksheet: Any, header_row: int, max_column: int, max_cell_chars: int) -> list[str]:
    rows = list(worksheet.iter_rows(min_row=header_row, max_row=header_row, max_col=max_column, values_only=True))
    raw_headers = rows[0] if rows else ()
    headers: list[str] = []
    used: set[str] = set()
    for index in range(max_column):
        header = _format_cell(raw_headers[index] if index < len(raw_headers) else None, max_cell_chars)
        if not header:
            header = f"未命名列_{get_column_letter(index + 1)}"
        original = header
        suffix = 2
        while header in used:
            header = f"{original}_{suffix}"
            suffix += 1
        used.add(header)
        headers.append(header)
    return headers


def _list_sheet_names(workbook_path: Path) -> list[str]:
    with zipfile.ZipFile(workbook_path) as archive:
        root = ET.fromstring(archive.read("xl/workbook.xml"))
    namespace = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    sheets_node = root.find("m:sheets", namespace)
    if sheets_node is None:
        return []
    return [str(sheet.attrib.get("name", "")) for sheet in sheets_node]


def _format_cell(value: Any, max_cell_chars: int) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        text = value.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(value, date):
        text = value.strftime("%Y-%m-%d")
    elif isinstance(value, time):
        text = value.strftime("%H:%M:%S")
    else:
        text = str(value)
    text = " ".join(text.replace("\t", " ").split())
    if len(text) > max_cell_chars:
        return text[: max_cell_chars - 1] + "…"
    return text
