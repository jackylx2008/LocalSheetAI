from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
import warnings
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

from localai.modules.config_loader import as_bool, as_int


@dataclass(frozen=True)
class ExcelSheetReadConfig:
    input_file: str
    sheet_name: str
    max_rows: int
    max_cols: int
    max_cell_chars: int
    max_total_chars: int
    data_only: bool = True

    @classmethod
    def from_config(cls, raw: dict[str, Any]) -> "ExcelSheetReadConfig":
        return cls(
            input_file=str(raw.get("input_file", "")).strip(),
            sheet_name=str(raw.get("sheet_name", "")).strip(),
            max_rows=as_int(raw.get("max_rows"), 80),
            max_cols=as_int(raw.get("max_cols"), 30),
            max_cell_chars=as_int(raw.get("max_cell_chars"), 200),
            max_total_chars=as_int(raw.get("max_total_chars"), 24000),
            data_only=as_bool(raw.get("data_only", True)),
        )


@dataclass(frozen=True)
class ExcelSheetSnapshot:
    workbook_path: Path
    requested_sheet_name: str
    actual_sheet_name: str
    available_sheet_names: list[str]
    max_row: int
    max_column: int
    rows_read: int
    cols_read: int
    truncated: bool
    text: str


@dataclass(frozen=True)
class ExcelSheetTarget:
    input_file: str
    sheet_name: str


def resolve_sheet_targets(config: ExcelSheetReadConfig, raw: dict[str, Any]) -> list[ExcelSheetTarget]:
    workbooks = raw.get("workbooks", {})
    if isinstance(workbooks, dict) and workbooks:
        targets: list[ExcelSheetTarget] = []
        for input_file, sheet_names in workbooks.items():
            if isinstance(sheet_names, str):
                names = _split_sheet_names(sheet_names)
            elif isinstance(sheet_names, list):
                names = [str(name).strip() for name in sheet_names if str(name).strip()]
            else:
                raise RuntimeError(f"Invalid sheet list for workbook {input_file}: {sheet_names!r}")
            if not names:
                raise RuntimeError(f"No sheet names configured for workbook: {input_file}")
            targets.extend(ExcelSheetTarget(input_file=str(input_file).strip(), sheet_name=name) for name in names)
        return targets

    return [ExcelSheetTarget(input_file=config.input_file, sheet_name=config.sheet_name)]


def find_excel_file(input_dir: Path, input_file: str = "") -> Path:
    if input_file:
        path = Path(input_file)
        if not path.is_absolute():
            path = input_dir / path
        if not path.exists():
            raise RuntimeError(f"Excel file not found: {path}")
        return path

    candidates = sorted(
        [path for path in input_dir.iterdir() if path.suffix.lower() in {".xlsx", ".xlsm"}],
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise RuntimeError(f"No .xlsx or .xlsm file found in input directory: {input_dir}")
    return candidates[0]


def list_sheet_names(workbook_path: Path) -> list[str]:
    with zipfile.ZipFile(workbook_path) as archive:
        root = ET.fromstring(archive.read("xl/workbook.xml"))
    namespace = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    sheets_node = root.find("m:sheets", namespace)
    if sheets_node is None:
        return []
    return [str(sheet.attrib.get("name", "")) for sheet in sheets_node]


def read_sheet_snapshot(workbook_path: Path, config: ExcelSheetReadConfig) -> ExcelSheetSnapshot:
    sheet_names = list_sheet_names(workbook_path)
    actual_sheet_name = resolve_sheet_name(config.sheet_name, sheet_names)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(
            workbook_path,
            read_only=True,
            data_only=config.data_only,
            keep_links=False,
        )
    try:
        worksheet = workbook[actual_sheet_name]
        max_rows = max(1, config.max_rows)
        max_cols = max(1, config.max_cols)
        max_cell_chars = max(1, config.max_cell_chars)
        max_total_chars = max(1, config.max_total_chars)

        lines: list[str] = []
        rows_read = 0
        cols_read = 0
        truncated = False
        total_chars = 0

        for row in worksheet.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True):
            values = [_format_cell(value, max_cell_chars) for value in row]
            trimmed_values = _trim_trailing_empty(values)
            if not trimmed_values:
                continue
            line = "\t".join(trimmed_values).rstrip()
            next_total = total_chars + len(line) + 1
            if next_total > max_total_chars:
                truncated = True
                break
            lines.append(line)
            total_chars = next_total
            rows_read += 1
            cols_read = max(cols_read, len(trimmed_values))

        if worksheet.max_row and worksheet.max_row > max_rows:
            truncated = True
        if worksheet.max_column and worksheet.max_column > max_cols:
            truncated = True

        return ExcelSheetSnapshot(
            workbook_path=workbook_path,
            requested_sheet_name=config.sheet_name,
            actual_sheet_name=actual_sheet_name,
            available_sheet_names=sheet_names,
            max_row=worksheet.max_row or 0,
            max_column=worksheet.max_column or 0,
            rows_read=rows_read,
            cols_read=cols_read,
            truncated=truncated,
            text="\n".join(lines),
        )
    finally:
        workbook.close()


def resolve_sheet_name(requested_sheet_name: str, available_sheet_names: list[str]) -> str:
    if not available_sheet_names:
        raise RuntimeError("Workbook has no sheets")
    if not requested_sheet_name:
        return available_sheet_names[0]
    if requested_sheet_name in available_sheet_names:
        return requested_sheet_name

    requested_normalized = requested_sheet_name.strip()
    for sheet_name in available_sheet_names:
        if sheet_name.strip() == requested_normalized:
            return sheet_name

    raise RuntimeError(
        f"Sheet not found: {requested_sheet_name}. Available sheets: {available_sheet_names}"
    )


def _format_cell(value: Any, max_cell_chars: int) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        text = value.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(value, date):
        text = value.strftime("%Y-%m-%d")
    elif isinstance(value, time):
        text = value.strftime("%H:%M:%S")
    else:
        text = str(value)
    text = " ".join(text.replace("\t", " ").split())
    if len(text) > max_cell_chars:
        return text[: max_cell_chars - 1] + "…"
    return text


def _trim_trailing_empty(values: list[str]) -> list[str]:
    end = len(values)
    while end > 0 and values[end - 1] == "":
        end -= 1
    return values[:end]


def _split_sheet_names(value: str) -> list[str]:
    return [item.strip() for item in value.split(",") if item.strip()]
